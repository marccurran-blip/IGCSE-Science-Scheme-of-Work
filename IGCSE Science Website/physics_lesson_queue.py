"""
Physics Lesson Queue Runner
============================
Reads the Physics sheet from SA_Science_Scheme_of_Work.xlsm one row at a time,
formats a Cowork prompt, copies it to your clipboard, and waits for you to
press Enter before loading the next lesson.

SETUP (run once):
    pip install openpyxl pyperclip

USAGE:
    python physics_lesson_queue.py

It remembers where you left off (saves progress to progress.txt in the same
folder). To restart from scratch, delete progress.txt.

To start from a specific lesson number, run:
    python physics_lesson_queue.py --start 15
"""

import os
import sys
import re
import openpyxl
import pyperclip

# ── CONFIGURATION ─────────────────────────────────────────────────────────
# Update this path to wherever your spreadsheet lives:
SPREADSHEET = "SA_Science_Scheme_of_Work.xlsm"
SHEET_NAME  = "Physics"
PROGRESS_FILE = "progress.txt"
# ──────────────────────────────────────────────────────────────────────────


def load_lessons(path, sheet):
    """Read every lesson row from the Physics sheet. Returns a list of dicts."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet]

    lessons = []
    for row in range(2, ws.max_row + 1):
        lesson_num_raw = ws.cell(row=row, column=3).value   # C - Lesson number
        lesson_name    = ws.cell(row=row, column=4).value   # D - Lesson name
        if not lesson_num_raw or not lesson_name:
            continue

        # Extract the number from "Lesson 1", "Lesson 67\n", etc.
        num_match = re.search(r"(\d+)", str(lesson_num_raw))
        if not num_match:
            continue
        lesson_num = int(num_match.group(1))

        lessons.append({
            "row":          row,
            "number":       lesson_num,
            "title":        str(lesson_name).strip(),
            "spec_points":  str(ws.cell(row=row, column=5).value or "").strip(),
            "requisitions": str(ws.cell(row=row, column=7).value or "").strip(),
            "planning":     str(ws.cell(row=row, column=9).value or "").strip(),
            "objectives":   str(ws.cell(row=row, column=10).value or "").strip(),
            "key_words":    str(ws.cell(row=row, column=11).value or "").strip(),
            "method":       str(ws.cell(row=row, column=12).value or "").strip(),
        })

    wb.close()
    return lessons


def build_prompt(lesson):
    """Format a lesson dict into the Cowork prompt."""

    # Work out practical info
    if lesson["requisitions"]:
        practical = lesson["requisitions"]
    elif lesson["method"]:
        practical = f"Practical: {lesson['method'][:200]}"
    else:
        practical = "No practical work"

    prompt = f"""Build the complete 7-file resource pack for Lesson {lesson['number']} - {lesson['title']} (IGCSE Physics).

## LESSON DATA

* Lesson {lesson['number']}: {lesson['title']}
* Spec points: {lesson['spec_points']}
* Learning objectives:
{lesson['objectives']}
* Key words:
{lesson['key_words']}
* {practical}

## ADDITIONAL CONTEXT

{lesson['planning']}

{f"Method details: {lesson['method']}" if lesson['method'] else ""}

## INSTRUCTIONS

Use the lesson-resources skill. This is IGCSE Physics. Use the following DARK BLUE colour palette (override the default Physics scheme):

```js
const C = {{primary:"0F172A", secondary:"1E3A5F", accent:"3B82F6", pale:"DBEAFE", cream:"EFF6FF", white:"FFFFFF", dark:"1B1B1E", grey:"6B7280", lightGrey:"F3F4F6", midGrey:"E5E7EB", amber:"F59E0B", green:"10B981", red:"EF4444", purple:"8B5CF6"}};
```
White text on primary background (#0F172A is dark midnight blue).
Read the scheme of work spreadsheet at /mnt/user-data/uploads/SA_Science_Scheme_of_Work.xlsm as the single source of truth (Physics sheet, row {lesson['row']}).
Build sequentially: assets then PPTX then DOCX then validate. Do NOT use parallel agents.
"""
    return prompt.strip()


def load_progress():
    """Return the last completed lesson number, or 0 if no progress file."""
    if os.path.exists(PROGRESS_FILE):
        with open(PROGRESS_FILE, "r") as f:
            content = f.read().strip()
            if content.isdigit():
                return int(content)
    return 0


def save_progress(lesson_num):
    """Save the last completed lesson number."""
    with open(PROGRESS_FILE, "w") as f:
        f.write(str(lesson_num))


def main():
    # Handle --start flag
    start_from = None
    if "--start" in sys.argv:
        idx = sys.argv.index("--start")
        if idx + 1 < len(sys.argv):
            start_from = int(sys.argv[idx + 1])

    # Check spreadsheet exists
    if not os.path.exists(SPREADSHEET):
        print(f"ERROR: Cannot find '{SPREADSHEET}'")
        print(f"  Put this script in the same folder as the spreadsheet,")
        print(f"  or update the SPREADSHEET path at the top of this file.")
        sys.exit(1)

    print("Loading spreadsheet...")
    lessons = load_lessons(SPREADSHEET, SHEET_NAME)
    print(f"Found {len(lessons)} Physics lessons.\n")

    # Sort by lesson number so we go in order
    lessons.sort(key=lambda x: x["number"])

    # Determine where to start
    if start_from is not None:
        last_done = start_from - 1
        print(f"Starting from lesson {start_from} (--start flag).\n")
    else:
        last_done = load_progress()
        if last_done > 0:
            print(f"Resuming after lesson {last_done} (from {PROGRESS_FILE}).")
            print(f"  Delete {PROGRESS_FILE} to restart from the beginning.\n")

    # Filter to remaining lessons
    remaining = [l for l in lessons if l["number"] > last_done]
    if not remaining:
        print("All lessons complete! Nothing left to do.")
        print(f"  Delete {PROGRESS_FILE} to start over.")
        return

    print(f"{len(remaining)} lessons remaining.\n")
    print("=" * 60)

    for i, lesson in enumerate(remaining):
        label = f"Lesson {lesson['number']} - {lesson['title']}"
        count = f"[{i+1}/{len(remaining)}]"

        print(f"\n{count}  {label}")
        print("-" * 60)

        # Build and copy the prompt
        prompt = build_prompt(lesson)
        try:
            pyperclip.copy(prompt)
            print("  Prompt copied to clipboard!")
        except Exception as e:
            print(f"  Could not copy to clipboard: {e}")
            print("  The prompt has been saved to _current_prompt.txt instead.")
            with open("_current_prompt.txt", "w", encoding="utf-8") as f:
                f.write(prompt)

        print(f"\n    Paste into Cowork now and wait for it to finish.")
        print(f"    Then come back here and press Enter.\n")

        try:
            input("    Press Enter when done (or Ctrl+C to stop)... ")
        except KeyboardInterrupt:
            print(f"\n\nStopped. Progress saved at lesson {lesson['number']}.")
            print(f"Run the script again to resume from here.")
            save_progress(last_done)  # save where we were before this one
            sys.exit(0)

        # Mark this lesson as done
        save_progress(lesson["number"])
        last_done = lesson["number"]
        print(f"    Lesson {lesson['number']} marked complete.\n")

    print("=" * 60)
    print("All Physics lessons complete!")
    print(f"Delete {PROGRESS_FILE} if you want to run again.")


if __name__ == "__main__":
    main()
