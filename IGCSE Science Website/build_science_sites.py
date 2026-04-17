#!/usr/bin/env python3
"""
Science Scheme of Work — Site Generator
Builds student and teacher HTML sites from the spreadsheet.
Run directly: python3 build_science_sites.py
Auto-update: scheduled task checks modification time before rebuilding.
"""

import openpyxl, re, json, sys, os, time
from pathlib import Path
from datetime import datetime

# ── Paths ────────────────────────────────────────────────────────────────────
# Use script's own location so this works across any Cowork session
BASE_DIR    = Path(__file__).resolve().parent
XLSX_PATH   = BASE_DIR / 'SA Science Scheme of Work.xlsm'
STUDENT_OUT = BASE_DIR / 'Science_Scheme_of_Work.html'
TEACHER_OUT = BASE_DIR / 'Science_Scheme_of_Work_TEACHER.html'
SA_SPEC_PATH = BASE_DIR / 'sa_specs_by_subject.json'

# ── Optional: Google Drive copy ───────────────────────────────────────────────
# If you have Google Drive for Desktop installed, set GDRIVE_DIR to your
# local Google Drive sync folder, e.g.:
#   GDRIVE_DIR = Path('C:/Users/marcc/Google Drive/My Drive/Science Resources')
# Leave as None to skip Google Drive copying.
GDRIVE_DIR = None

# Fallback SA specs (extracted from 4SS0 PDF) — stored inline so script is self-contained
SA_SPECS_INLINE = {
  "Biology": ["1.1","1.2","1.3","1.4","2.1","2.2","2.3","2.4","2.7","2.8","2.9","2.10","2.11","2.12","2.13","2.15","2.16","2.18","2.19","2.20","2.21","2.23","2.27","2.29","2.34","2.35","2.36","2.37","2.38","2.46","2.47","2.48","2.51","2.52","2.59","2.60","2.61","2.62","2.65","2.68","2.69","3.1","3.2","3.3","3.4","3.8","3.13","3.15","3.19","3.20","3.23","3.25","3.26","3.27","3.31","3.33","3.34","3.38","4.1","4.2","4.5","4.6","4.7","4.8","4.9","4.10","5.1","5.2","5.5","5.6","5.12","5.13","5.14","5.15","5.16"],
  "Chemistry": ["1.1","1.2","1.3","1.8","1.9","1.10","1.11","1.12","1.13","1.14","1.15","1.16","1.17","1.18","1.21","1.25","1.26","1.37","1.38","1.39","1.41","1.42","1.44","1.47","1.49","2.1","2.2","2.3","2.5","2.6","2.9","2.10","2.11","2.13","2.14","2.15","2.17","2.18","2.19","2.28","2.29","2.30","2.31","2.32","2.44","2.45","2.46","2.48","2.49","3.1","3.2","3.3","3.8","3.9","3.10","3.12","3.15","4.1","4.2","4.7","4.9","4.10","4.11","4.12","4.13","4.14","4.15","4.16","4.19","4.20","4.21","4.23","4.24","4.25","4.26","4.28","4.44","4.45","4.46","4.47"],
  "Physics": ["1.1","1.3","1.4","1.5","1.6","1.7","1.8","1.9","1.11","1.12","1.16","1.17","1.18","1.19","1.20","2.1","2.4","2.6","2.8","2.9","2.10","2.12","2.13","2.14","2.16","2.19","3.1","3.3","3.4","3.5","3.7","3.9","3.10","3.11","3.12","3.13","3.14","3.15","3.17","3.20","3.21","3.23","4.1","4.2","4.3","4.4","4.5","4.11","4.12","4.13","4.14","4.15","4.16","4.17","5.1","5.5","5.6","5.15","5.16","5.17","5.18","5.19","5.20","6.1","6.4","6.6","6.7","6.8","6.12","6.13","6.14","7.1","7.2","7.3","7.4","7.5","7.6","7.10","7.12","7.14","7.15","7.16","7.17","7.18","7.19","7.22","7.25","8.1","8.2","8.3","8.4","8.5","8.7","8.8","8.9"]
}

# ── Data extraction ───────────────────────────────────────────────────────────
def extract_hyperlink(cell):
    val = cell.value or ''
    if isinstance(val, str) and val.startswith('=HYPERLINK('):
        m = re.match(r'=HYPERLINK\("([^"]+)",\s*"([^"]+)"\)', val)
        if m:
            return m.group(1), m.group(2).strip()
    url = cell.hyperlink.target if cell.hyperlink else None
    return url, (str(val).strip() if val else None)

# Matches bare URLs in plain text lines
_URL_RE = re.compile(r'https?://[^\s"\'<>]+')

def extract_planning_items(cell):
    """Return a list of {'text': str, 'url': str|None} dicts for the Planning cell.

    Handles three cases:
      1. =HYPERLINK("url","label") formula  → single link item
      2. Cell with a .hyperlink attribute   → single link item (multi-line text kept)
      3. Plain text (possibly multi-line)   → one item per line; bare URLs auto-linked
    """
    if cell is None:
        return []
    val = cell.value
    if not val:
        return []

    # Case 1: =HYPERLINK formula
    if isinstance(val, str) and val.startswith('=HYPERLINK('):
        m = re.match(r'=HYPERLINK\("([^"]+)",\s*"([^"]+)"\)', val)
        if m:
            return [{'url': m.group(1), 'text': m.group(2).strip()}]
        return []   # malformed — skip

    # Case 2: traditional Excel hyperlink (<hyperlink> XML element)
    hyper_url = cell.hyperlink.target if cell.hyperlink else None
    text = str(val).strip()
    if not text:
        return []

    if hyper_url:
        # Entire cell is one link; display text may be multi-line notes.
        # Render as a "Open folder" button at top + plain text lines below.
        lines = [l.strip() for l in text.split('\n') if l.strip()]
        items = [{'url': hyper_url, 'text': '&#128193; Open lesson folder', 'folder_link': True}]
        items.extend({'url': None, 'text': l} for l in lines)
        return items

    # Case 3: plain text — split into lines, auto-link bare URLs
    lines = [l.strip() for l in text.split('\n') if l.strip()]
    items = []
    for line in lines:
        m = _URL_RE.search(line)
        if m:
            items.append({'url': m.group(0), 'text': line})
        else:
            items.append({'url': None, 'text': line})
    return items

def parse_spec_points(raw_text):
    if not raw_text:
        return []
    lines   = raw_text.strip().split('\n')
    points  = []
    cur_num = cur_let = None
    cur_lines = []
    spec_re = re.compile(r'^(\d+\.\d+)([BCP]?)\s+(.*)')
    for line in lines:
        line = line.strip()
        m = spec_re.match(line)
        if m:
            if cur_num:
                points.append({'number': cur_num, 'letter': cur_let,
                                'text': ' '.join(cur_lines).strip(),
                                'is_single_only': cur_let in ('B','C','P')})
            cur_num, cur_let = m.group(1), m.group(2)
            cur_lines = [m.group(3).strip()]
        elif cur_num and line:
            cur_lines.append(line)
    if cur_num:
        points.append({'number': cur_num, 'letter': cur_let,
                       'text': ' '.join(cur_lines).strip(),
                       'is_single_only': cur_let in ('B','C','P')})
    return points

def load_data():
    wb = openpyxl.load_workbook(str(XLSX_PATH), keep_vba=True)
    all_data = {}
    for sheet_name in ['Biology', 'Chemistry', 'Physics']:
        ws      = wb[sheet_name]
        lessons = []
        cur_year = cur_term = None
        for row in ws.iter_rows(min_row=2):
            cells = {c.column: c for c in row if hasattr(c, 'value')}
            a = cells.get(1); b = cells.get(2); c3 = cells.get(3)
            d = cells.get(4); e = cells.get(5); f = cells.get(6)
            g = cells.get(7); h = cells.get(8); i = cells.get(9)
            j = cells.get(10); k = cells.get(11); l_cell = cells.get(12)
            if a and a.value: cur_year = str(a.value).strip()
            if b and b.value: cur_term = str(b.value).strip()
            if not c3: continue
            url, label = extract_hyperlink(c3)
            if not url: continue
            name         = (d.value or '') if d else ''
            spec_raw     = (e.value or '') if e else ''
            pages        = (f.value or '') if f else ''
            req_raw       = (g.value or '') if g else ''
            notes_raw     = (h.value or '') if h else ''
            planning_items = extract_planning_items(i) if i else []
            objectives_raw = (j.value or '') if j else ''
            keywords_raw   = (k.value or '') if k else ''
            try:
                method_raw = str(l_cell.value or '').strip() if l_cell else ''
            except:
                method_raw = ''
            name = str(name).strip()
            if not name: continue
            lessons.append({
                'year':          cur_year or '',
                'term':          cur_term or '',
                'lesson_label':  label,
                'url':           url,
                'lesson_name':   name,
                'spec_points':   parse_spec_points(str(spec_raw).strip()),
                'pages':         str(pages).strip(),
                'requisitions':  str(req_raw).strip(),
                'notes':         str(notes_raw).strip(),
                'planning':      planning_items,
                'objectives':    str(objectives_raw).strip(),
                'keywords':      str(keywords_raw).strip(),
                'method':        method_raw,
            })
        all_data[sheet_name] = lessons

    # ── Load assessment dates from the two test tabs ──
    def _read_test_columns(ws, name_col, start_col, end_col):
        from datetime import timedelta
        tests = []
        for row_num in range(3, 20):
            name_cell = ws.cell(row=row_num, column=name_col).value
            start_cell = ws.cell(row=row_num, column=start_col).value
            end_cell = ws.cell(row=row_num, column=end_col).value
            if not name_cell or not start_cell:
                break
            # If end cell is a formula string (not cached) but start is a date,
            # fall back to start + 7 days as a reasonable estimate
            if end_cell is None or (isinstance(end_cell, str) and end_cell.startswith('=')):
                end_cell = start_cell + timedelta(days=7) if hasattr(start_cell, 'strftime') else None
            tests.append({
                'name': str(name_cell).strip(),
                'start': start_cell.strftime('%d %b %Y') if hasattr(start_cell, 'strftime') else str(start_cell)[:10],
                'start_iso': start_cell.strftime('%Y-%m-%d') if hasattr(start_cell, 'strftime') else '',
                'end': end_cell.strftime('%d %b %Y') if hasattr(end_cell, 'strftime') else str(end_cell or '')[:10],
                'end_iso': end_cell.strftime('%Y-%m-%d') if hasattr(end_cell, 'strftime') else '',
            })
        # Tag tests by year group: first 4 = Year 10, remainder = Year 11
        # (JS will highlight the next upcoming test independently within each group)
        for i, t in enumerate(tests):
            t['year_group'] = 'y10' if i < 4 else 'y11'
        return tests

    assessment_dates = {}
    # Tab 1: Single-subject tests — Bio cols 2-4, Chem cols 7-9, Phys cols 12-14
    tab1 = 'Bio, chem, phys test'
    if tab1 in wb.sheetnames:
        ws = wb[tab1]
        assessment_dates['Biology'] = _read_test_columns(ws, 2, 3, 4)
        assessment_dates['Chemistry'] = _read_test_columns(ws, 7, 8, 9)
        assessment_dates['Physics'] = _read_test_columns(ws, 12, 13, 14)
    # Tab 2: Double & Single Award tests — Bio cols 1-3, Chem cols 6-8, Phys cols 11-13
    tab2 = 'Double and Single Award Test'
    if tab2 in wb.sheetnames:
        ws = wb[tab2]
        assessment_dates['DS_Biology'] = _read_test_columns(ws, 1, 2, 3)
        assessment_dates['DS_Chemistry'] = _read_test_columns(ws, 6, 7, 8)
        assessment_dates['DS_Physics'] = _read_test_columns(ws, 11, 12, 13)
    all_data['_assessments'] = assessment_dates

    return all_data

# ── Card rendering ────────────────────────────────────────────────────────────
import re as _re

_SUPER = str.maketrans('0123456789+-', '⁰¹²³⁴⁵⁶⁷⁸⁹⁺⁻')

def _fix_science_text(s):
    """Convert scientific notation to proper Unicode: cm3→cm³, Na+→Na⁺, CO2→CO₂, etc."""
    s = str(s)
    # ── Units with powers: cm3→cm³, dm3→dm³, m2→m², etc. ──
    s = _re.sub(r'\b(cm|dm|mm|km)([23])\b', lambda m: m.group(1) + m.group(2).translate(_SUPER), s)
    s = _re.sub(r'\b(m)([23])\b(?!\w)', lambda m: m.group(1) + m.group(2).translate(_SUPER), s)
    # ── Polyatomic ion charges: NH4+→NH₄⁺, CO32-→CO₃²⁻, SO42-→SO₄²⁻, NO3-→NO₃⁻ ──
    s = _re.sub(r'\bNH4\+', 'NH₄⁺', s)
    s = _re.sub(r'\bNH4\b', 'NH₄', s)
    s = _re.sub(r'\bCO32[\-−]', 'CO₃²⁻', s)
    s = _re.sub(r'\bSO42[\-−]', 'SO₄²⁻', s)
    s = _re.sub(r'\bNO3[\-−]', 'NO₃⁻', s)
    s = _re.sub(r'\bOH[\-−](?![a-zA-Z])', 'OH⁻', s)
    s = _re.sub(r'\bMnO4[\-−]', 'MnO₄⁻', s)
    s = _re.sub(r'\bCr2O72[\-−]', 'Cr₂O₇²⁻', s)
    s = _re.sub(r'\bPO43[\-−]', 'PO₄³⁻', s)
    # ── Simple ion charges (only known ions to avoid false positives on URLs) ──
    # Cations
    for ion in ['H', 'Na', 'K', 'Li', 'Ag', 'Cu', 'Zn', 'Fe', 'Ca', 'Mg', 'Ba', 'Al', 'Pb', 'Mn', 'Cr', 'Ni', 'Co', 'Sn']:
        s = _re.sub(rf'\b{ion}\+(?!\w)', f'{ion}⁺', s)
        s = _re.sub(rf'\b{ion}2\+(?!\w)', f'{ion}²⁺', s)
        s = _re.sub(rf'\b{ion}3\+(?!\w)', f'{ion}³⁺', s)
    # Anions
    for ion in ['Cl', 'Br', 'I', 'F', 'O', 'S']:
        s = _re.sub(rf'\b{ion}[\-−](?![a-zA-Z])', f'{ion}⁻', s)
        s = _re.sub(rf'\b{ion}2[\-−](?![a-zA-Z])', f'{ion}²⁻', s)
    # ── Chemical subscripts: CO2→CO₂, H2O→H₂O, etc. ──
    s = _re.sub(r'\bCO2\b', 'CO₂', s)
    s = _re.sub(r'\bH2O\b', 'H₂O', s)
    s = _re.sub(r'\bH2SO4\b', 'H₂SO₄', s)
    s = _re.sub(r'\bH2S\b', 'H₂S', s)
    s = _re.sub(r'\bO2\b', 'O₂', s)
    s = _re.sub(r'\bN2\b', 'N₂', s)
    s = _re.sub(r'\bCl2\b', 'Cl₂', s)
    s = _re.sub(r'\bBr2\b', 'Br₂', s)
    s = _re.sub(r'\bI2\b', 'I₂', s)
    s = _re.sub(r'\bCaCO3\b', 'CaCO₃', s)
    s = _re.sub(r'\bNa2CO3\b', 'Na₂CO₃', s)
    s = _re.sub(r'\bCaCl2\b', 'CaCl₂', s)
    s = _re.sub(r'\bMgCl2\b', 'MgCl₂', s)
    s = _re.sub(r'\bFeCl2\b', 'FeCl₂', s)
    s = _re.sub(r'\bFeCl3\b', 'FeCl₃', s)
    s = _re.sub(r'\bNaOH\b', 'NaOH', s)
    s = _re.sub(r'\bHCl\b', 'HCl', s)
    s = _re.sub(r'\bHNO3\b', 'HNO₃', s)
    s = _re.sub(r'\bNaCl\b', 'NaCl', s)
    return s

def esc(s):
    """HTML-escape text content with science formatting."""
    return _fix_science_text(str(s)).replace('&','&amp;').replace('<','&lt;').replace('>','&gt;')

def esc_url(s):
    """HTML-escape for URL attributes — NO science formatting."""
    return str(s).replace('&','&amp;').replace('<','&lt;').replace('>','&gt;')

def attr_val(s):
    """Escape a string for safe embedding inside an HTML attribute value."""
    return _fix_science_text(s).replace('&','&amp;').replace('<','&lt;').replace('>','&gt;').replace('"','&quot;')

def render_teacher_section(icon, title, css_class, raw, line_class):
    """Collapsible teacher section with copy button. Always shown; 'None' when cell empty."""
    if not raw or raw in ('nan', ''):
        plain_text  = 'None'
        content_html = f'<div class="{line_class} empty-cell">None</div>'
    else:
        lines        = [l.strip() for l in raw.strip().split('\n') if l.strip()]
        plain_text   = '\n'.join(lines)
        content_html = ''.join(f'<div class="{line_class}">{esc(l)}</div>' for l in lines)

    safe_text = attr_val(plain_text)
    return (
        f'<div class="teacher-section {css_class}">'
        f'<div class="tsec-header" onclick="toggleSection(this)">'
        f'<span class="tsec-icon">{icon}</span>'
        f'<span class="tsec-title">{title}</span>'
        f'<button class="copy-btn" onclick="copySection(event,this)" '
        f'data-plain="{safe_text}" title="Copy to clipboard">&#128203; Copy</button>'
        f'<span class="tsec-arrow">&#9654;</span>'
        f'</div>'
        f'<div class="tsec-body collapsed">{content_html}</div>'
        f'</div>'
    )

def render_requisitions(raw, lesson_label='', lesson_name=''):
    """Render Requisitions with lesson number and name prepended."""
    if lesson_label or lesson_name:
        header_line = f'{lesson_label} — {lesson_name}'.strip(' —')
        if raw and raw not in ('nan', ''):
            raw = f'{header_line}\n{raw}'
        else:
            raw = header_line
    return render_teacher_section('&#128101;', 'Requisitions', 'req-section', raw, 'req-line')

def render_notes(raw):
    return render_teacher_section('&#9888;&#65039;', 'Notes &amp; Safety', 'notes-safety-section', raw, 'notes-line')

def render_method(raw):
    """Render Method Details accordion — only shown if content exists."""
    if not raw or str(raw).strip() in ('nan', '', 'None'):
        return ''
    lines = [l.strip() for l in str(raw).strip().split('\n') if l.strip()]
    if not lines:
        return ''
    plain_text = '\n'.join(lines)
    content_html = ''.join(f'<div class="notes-line">{esc(l)}</div>' for l in lines)
    safe_text = attr_val(plain_text)
    return (
        f'<div class="teacher-section method-section">'
        f'<div class="tsec-header" onclick="toggleSection(this)">'
        f'<span class="tsec-icon">&#9878;</span>'
        f'<span class="tsec-title">Method Details</span>'
        f'<button class="copy-btn" onclick="copySection(event,this)" '
        f'data-plain="{safe_text}" title="Copy to clipboard">&#128203; Copy</button>'
        f'<span class="tsec-arrow">&#9654;</span>'
        f'</div>'
        f'<div class="tsec-body collapsed">{content_html}</div>'
        f'</div>'
    )

def render_planning(items):
    """Render Planning accordion; items is a list of {'text', 'url'} dicts."""
    icon = '&#128196;'
    title = 'Planning'

    if not items:
        plain_text   = 'None'
        html_copy    = '<p><em>None</em></p>'
        content_html = '<div class="notes-line empty-cell">None</div>'
    else:
        plain_lines   = []
        html_parts    = []   # rich-text HTML for clipboard
        content_parts = []   # display HTML inside the accordion
        for item in items:
            t = item['text']
            u = item.get('url')
            if item.get('folder_link') and u:
                # Plain copy: just the URL; Rich copy: clickable hyperlink
                plain_lines.append(u)
                html_parts.append(f'<p><a href="{u}">{u}</a></p>')
                content_parts.append(
                    f'<a href="{esc_url(u)}" target="_blank" class="folder-link-btn">'
                    f'{t} &#8594;</a>'
                )
            elif u:
                plain_lines.append(f'{t}  ({u})')
                html_parts.append(f'<p><a href="{u}">{esc(t)}</a></p>')
                content_parts.append(
                    f'<div class="notes-line">'
                    f'<a href="{esc_url(u)}" target="_blank" class="plan-link">'
                    f'{esc(t)}</a></div>'
                )
            else:
                plain_lines.append(t)
                html_parts.append(f'<p>{esc(t)}</p>')
                # Detect section headers (e.g. "Specification Points:", "Requisitions:") to add spacing
                is_lesson_title = t.startswith('Lesson ') or t.startswith('Test ')
                is_section_head = t.rstrip().endswith(':') and len(t) < 60
                if is_lesson_title:
                    cls = 'plan-lesson-title'
                elif is_section_head:
                    cls = 'plan-section-header'
                else:
                    cls = 'notes-line'
                content_parts.append(f'<div class="{cls}">{esc(t)}</div>')
        plain_text   = '\n'.join(plain_lines)
        html_copy    = ''.join(html_parts)
        content_html = ''.join(content_parts)

    safe_text    = attr_val(plain_text)
    # Escape html_copy for embedding in an attribute; browser will decode it back for JS
    safe_html    = attr_val(html_copy)
    return (
        f'<div class="teacher-section notes-section">'
        f'<div class="tsec-header" onclick="toggleSection(this)">'
        f'<span class="tsec-icon">{icon}</span>'
        f'<span class="tsec-title">{title}</span>'
        f'<button class="copy-btn" onclick="copySection(event,this)" '
        f'data-plain="{safe_text}" data-copyhtml="{safe_html}" '
        f'title="Copy to clipboard">&#128203; Copy</button>'
        f'<span class="tsec-arrow">&#9654;</span>'
        f'</div>'
        f'<div class="tsec-body collapsed">{content_html}</div>'
        f'</div>'
    )

def render_card(lesson, subject, spec_points_override=None, teacher=False):
    sl   = subject.lower()
    label = esc(lesson['lesson_label'])
    name  = esc(lesson['lesson_name'])
    url   = lesson['url']
    pages = lesson['pages']
    specs = spec_points_override if spec_points_override is not None else lesson['spec_points']
    year  = lesson['year']
    term  = lesson['term']
    year_data = year.replace(' ','').lower()
    term_data = term.replace(' ','').lower()

    pages_html = ''
    if pages and pages not in ('nan',''):
        pages_html = f'<div class="card-pages"><span class="pages-icon">&#128214;</span> Textbook: {esc(pages)}</div>'

    spec_items = ''
    for sp in specs:
        num  = sp['number'] + sp['letter']
        txt  = esc(sp['text'])
        badge = ''
        if sp['letter'] == 'B': badge = '<span class="badge badge-bio">Bio only</span>'
        elif sp['letter'] == 'C': badge = '<span class="badge badge-chem">Chem only</span>'
        elif sp['letter'] == 'P': badge = '<span class="badge badge-phys">Phys only</span>'
        spec_items += (f'<li class="spec-item">'
                       f'<span class="spec-num">{num}</span> '
                       f'<span class="spec-text">{txt}</span> {badge}</li>\n')
    specs_html = (f'<div class="student-section student-specs">'
                  f'<div class="student-section-title">&#128218; Specification Points</div>'
                  f'<ul class="spec-list">{spec_items}</ul></div>') if spec_items else ''

    yr_badge = (f'<span class="year-badge">{esc(year)} &middot; {esc(term)}</span>'
                if year and term else '')

    # Notes / safety HTML (shared by both views)
    notes_raw = lesson.get('notes', '')
    notes_clean = notes_raw if notes_raw and notes_raw not in ('nan', 'None', 'N/A') else ''

    # --- Shared: folder button + objectives + keywords HTML ---
    folder_url = next(
        (item['url'] for item in lesson['planning'] if item.get('folder_link') and item.get('url')),
        None
    )
    folder_btn = (
        f'<a href="{esc_url(folder_url)}" target="_blank" class="student-folder-btn">'
        f'&#128193; Open Lesson Folder &#8594;</a>'
    ) if folder_url else ''

    obj_raw = lesson.get('objectives', '')
    obj_html = ''
    if obj_raw and obj_raw not in ('nan', 'None', 'N/A', ''):
        obj_lines = [l.strip().lstrip('• ').strip() for l in obj_raw.split('\n') if l.strip()]
        if obj_lines:
            items = ''.join(f'<li>{esc(l)}</li>' for l in obj_lines)
            obj_html = f'<div class="student-section student-objectives"><div class="student-section-title">&#127919; Learning Objectives</div><ul>{items}</ul></div>'

    kw_raw = lesson.get('keywords', '')
    kw_html = ''
    if kw_raw and kw_raw not in ('nan', 'None', 'N/A', ''):
        kw_lines = [l.strip().lstrip('• ').strip() for l in kw_raw.split('\n') if l.strip()]
        if kw_lines:
            items = ''.join(f'<li>{esc(l)}</li>' for l in kw_lines)
            kw_html = f'<div class="student-section student-keywords"><div class="student-section-title">&#128273; Key Words</div><ul>{items}</ul></div>'

    # Safety notes as a visual box
    safety_html = ''
    if notes_clean:
        safety_lines = [l.strip() for l in notes_clean.split('\n') if l.strip()]
        if safety_lines:
            safety_items = ''.join(f'<li>{esc(l)}</li>' for l in safety_lines)
            safety_html = f'<div class="student-section student-safety"><div class="student-section-title">&#9888;&#65039; Notes &amp; Safety</div><ul>{safety_items}</ul></div>'

    # Single scrollable area: specs → objectives → keywords → safety
    scroll_content = f'{specs_html}{obj_html}{kw_html}{safety_html}'
    scroll_div = f'<div class="info-scroll">{scroll_content}</div>' if scroll_content else ''

    if teacher:
        method_raw  = lesson.get('method', '')
        bottom_html = (
            f'<div class="card-bottom">'
            f'{render_method(method_raw)}'
            f'{render_requisitions(lesson["requisitions"], lesson.get("lesson_label",""), lesson.get("lesson_name",""))}'
            f'{render_planning(lesson["planning"])}'
            f'</div>'
        )
        top_html   = f'<div class="card-top">{folder_btn}{pages_html}{scroll_div}</div>'
        card_body  = f'<div class="card-body">{top_html}{bottom_html}</div>'
    else:
        top_html  = f'<div class="card-top">{folder_btn}{pages_html}{scroll_div}</div>'
        card_body = (
            f'<div class="card-body">'
            f'{top_html}'
            f'</div>'
        )

    return (f'<div class="lesson-card" data-subject="{sl}" data-year="{year_data}" data-term="{term_data}">'
            f'<div class="card-header card-header-{sl}">'
            f'<div class="card-title-row">'
            f'<a href="{url}" target="_blank" class="lesson-link">'
            f'<span class="lesson-label">{label}</span>'
            f'<span class="lesson-name">{name}</span>'
            f'<span class="link-arrow">&#8594;</span></a>'
            f'{yr_badge}</div></div>'
            f'{card_body}</div>')

# ── Filter helpers ─────────────────────────────────────────────────────────────
def filter_double(lesson):
    relevant = [p for p in lesson['spec_points'] if not p['is_single_only']]
    if not relevant and lesson['spec_points']: return None
    return relevant

def filter_sa(lesson, subject, sa_sets):
    sa_set = sa_sets[subject]
    relevant = [p for p in lesson['spec_points']
                if not p['is_single_only'] and p['number'] in sa_set]
    if not relevant and lesson['spec_points']: return None
    return relevant

# ── CSS ────────────────────────────────────────────────────────────────────────
def get_css(teacher=False):
    teacher_extra = """
/* ── Teacher view: uniform card heights, scroll inside card, full content in modal ── */
.lesson-card { height: 100%; }
.card-body { flex: 1; overflow: hidden; display: flex; flex-direction: column; }
.card-top { flex: 1 1 auto; overflow-y: auto; min-height: 0; padding-bottom: 4px; }
.card-bottom { max-height: 200px; overflow-y: auto; }
.card-top::-webkit-scrollbar { width: 4px; }
.card-top::-webkit-scrollbar-thumb { background: #ccc; border-radius: 2px; }
.card-bottom {
  flex-shrink: 0;
  border-top: 1px solid var(--border);
  padding-top: 10px;
  margin-top: 10px;
}
.spec-list { padding-right: 4px; }

/* ── Teacher sections ── */
.teacher-section { margin-top:10px; border-radius:10px; overflow:hidden; }

/* Accordion header */
.tsec-header {
  display:flex; align-items:center; gap:8px;
  padding:9px 14px; color:white; cursor:pointer;
  user-select:none; transition: filter .15s;
}
.tsec-header:hover { filter:brightness(1.1); }
.req-section   { border:1px solid #90caf9; }
.req-section   .tsec-header { background:linear-gradient(135deg,#1565c0 0%,#1e88e5 100%); }
.notes-safety-section { border:1px solid #ef9a9a; }
.notes-safety-section .tsec-header { background:linear-gradient(135deg,#c62828 0%,#e53935 100%); }
.notes-section { border:1px solid #b39ddb; }
.notes-section .tsec-header { background:linear-gradient(135deg,#4527a0 0%,#7e57c2 100%); }
.method-section { border:1px solid #a5d6a7; }
.method-section .tsec-header { background:linear-gradient(135deg,#2e7d32 0%,#43a047 100%); }
.method-section .tsec-body { background:#e8f5e9; }
.tsec-icon  { font-size:14px; flex-shrink:0; }
.tsec-title { font-size:11px; font-weight:700; text-transform:uppercase; letter-spacing:.7px; flex:1; }
.tsec-arrow {
  font-size:10px; flex-shrink:0; transition:transform .22s ease;
  color:rgba(255,255,255,.7);
}
.tsec-header.open .tsec-arrow { transform: rotate(90deg); }

/* Copy button */
.copy-btn {
  margin-left:auto; margin-right:6px;
  padding:2px 9px; border-radius:10px; border:1px solid rgba(255,255,255,.35);
  background:rgba(255,255,255,.18); color:white;
  font-size:11px; font-weight:600; cursor:pointer;
  transition: background .15s; white-space:nowrap; flex-shrink:0;
}
.copy-btn:hover { background:rgba(255,255,255,.32); }
.copy-btn.copied { background:rgba(100,220,130,.4); border-color:rgba(100,220,130,.6); }

/* Collapsible body */
.tsec-body {
  max-height: 0; overflow: hidden;
  transition: max-height .28s ease, padding .28s ease;
  padding: 0 14px; font-size:12.5px; line-height:1.6;
}
.tsec-body:not(.collapsed) {
  max-height: 340px;
  overflow-y: auto;
  padding: 10px 14px;
}
.tsec-body::-webkit-scrollbar { width: 5px; }
.tsec-body::-webkit-scrollbar-thumb { background:#bbb; border-radius:4px; }
.tsec-body::-webkit-scrollbar-thumb:hover { background:#999; }
.req-section .tsec-body { background:#e3f2fd; }
.notes-safety-section .tsec-body { background:#fce4ec; }
.notes-section .tsec-body { background:#ede7f6; }
.req-line, .notes-line { color:var(--text-mid); padding:2px 0; }
.req-line:not(:last-child), .notes-line:not(:last-child) { border-bottom:1px dotted #eee; }
.empty-cell { color:#bbb !important; font-style:italic; }
.plan-link { color:#1565c0; text-decoration:underline; text-underline-offset:2px; word-break:break-all; }
.plan-lesson-title { color:var(--text-dark); padding:2px 0 4px; font-weight:700; font-size:13px; border-bottom:none !important; }
.plan-section-header { color:var(--text-dark); padding:6px 0 2px; margin-top:8px; font-weight:700; font-size:12px; border-bottom:none !important; border-top:1px solid #ccc; }
.plan-link:hover { color:#0d47a1; }
.folder-link-btn {
  display:inline-flex; align-items:center; gap:5px;
  margin:6px 0 8px; padding:5px 12px;
  background:#e8f0fe; color:#1565c0; border:1px solid #bbdefb;
  border-radius:6px; font-size:12px; font-weight:600;
  text-decoration:none; transition:background .15s;
}
.folder-link-btn:hover { background:#bbdefb; color:#0d47a1; }
""" if teacher else ""

    return f"""
:root {{
  --bio-color:#1b5e35;--bio-light:#d8f3dc;--bio-mid:#40916c;
  --chem-color:#7c2f00;--chem-light:#fdebd9;--chem-mid:#d4570a;
  --phys-color:#0d3d6e;--phys-light:#ddeeff;--phys-mid:#1976d2;
  --double-color:#4a4a6a;--double-light:#eeeef8;--double-mid:#6b6b9a;
  --sa-color:#6b3800;--sa-light:#fff3e0;--sa-mid:#f57c00;
  --text-dark:#1a1a2e;--text-mid:#444;--text-light:#777;
  --border:#e0e0e0;--bg:#f4f6fb;--card-bg:#fff;--shadow:0 2px 12px rgba(0,0,0,.08);
}}
*{{box-sizing:border-box;margin:0;padding:0;}}
body{{font-family:'Segoe UI',system-ui,sans-serif;background:var(--bg);color:var(--text-dark);min-height:100vh;}}
.site-header{{background:linear-gradient(135deg,#1a1a2e 0%,#16213e 50%,#0f3460 100%);color:white;padding:36px 24px 28px;text-align:center;position:relative;overflow:visible;}}
.site-header::before{{content:'';position:absolute;inset:0;background:radial-gradient(ellipse at 20% 50%,rgba(64,145,108,.18) 0%,transparent 55%),radial-gradient(ellipse at 80% 50%,rgba(25,118,210,.15) 0%,transparent 55%);pointer-events:none;overflow:hidden;}}
.header-eyebrow{{display:inline-block;background:rgba(255,255,255,.12);border:1px solid rgba(255,255,255,.25);border-radius:20px;padding:5px 16px;font-size:12px;font-weight:700;letter-spacing:1.5px;text-transform:uppercase;margin-bottom:14px;position:relative;}}
.teacher-banner{{display:inline-block;background:rgba(255,200,50,.2);border:1px solid rgba(255,200,50,.4);border-radius:8px;padding:4px 14px;font-size:12px;font-weight:700;letter-spacing:.5px;margin-left:8px;color:#ffe082;}}
.edition-link{{display:inline-block;background:rgba(255,255,255,.12);border:1px solid rgba(255,255,255,.3);border-radius:8px;padding:4px 14px;font-size:12px;font-weight:700;letter-spacing:.5px;margin-left:8px;color:rgba(255,255,255,.85);text-decoration:none;transition:background .2s,color .2s;}}
.edition-link:hover{{background:rgba(255,255,255,.25);color:white;}}
.header-title{{font-size:clamp(24px,4vw,40px);font-weight:800;letter-spacing:-.5px;margin-bottom:8px;position:relative;}}
.header-subtitle{{font-size:14px;color:rgba(255,255,255,.65);position:relative;}}
.spec-codes{{display:flex;gap:10px;justify-content:center;flex-wrap:wrap;margin-top:18px;position:relative;}}
.spec-tag{{background:rgba(255,255,255,.1);border:1px solid rgba(255,255,255,.2);border-radius:8px;padding:4px 12px;font-size:12px;font-weight:600;letter-spacing:.5px;color:white;text-decoration:none;transition:background .2s,border-color .2s;}}
a.spec-tag:hover{{background:rgba(255,255,255,.22);border-color:rgba(255,255,255,.45);}}
.nav-wrapper{{background:white;border-bottom:1px solid var(--border);position:sticky;top:0;z-index:100;box-shadow:0 2px 8px rgba(0,0,0,.05);}}
.nav-tabs{{display:flex;overflow-x:auto;max-width:1200px;margin:0 auto;padding:0 16px;scrollbar-width:none;}}
.nav-tabs::-webkit-scrollbar{{display:none;}}
.nav-tab{{flex-shrink:0;padding:15px 20px;font-size:13.5px;font-weight:600;cursor:pointer;border:none;background:transparent;color:var(--text-light);border-bottom:3px solid transparent;transition:all .2s;white-space:nowrap;display:flex;align-items:center;gap:7px;}}
.nav-tab:hover{{color:var(--text-dark);}}
.nav-tab[data-tab="biology"].active{{border-color:var(--bio-mid);color:var(--bio-color);}}
.nav-tab[data-tab="chemistry"].active{{border-color:var(--chem-mid);color:var(--chem-color);}}
.nav-tab[data-tab="physics"].active{{border-color:var(--phys-mid);color:var(--phys-color);}}
.nav-tab[data-tab="double"].active{{border-color:var(--double-mid);color:var(--double-color);}}
.nav-tab[data-tab="sa"].active{{border-color:var(--sa-mid);color:var(--sa-color);}}
.tab-dot{{width:8px;height:8px;border-radius:50%;display:inline-block;flex-shrink:0;}}
[data-tab="biology"] .tab-dot{{background:var(--bio-mid);}}
[data-tab="chemistry"] .tab-dot{{background:var(--chem-mid);}}
[data-tab="physics"] .tab-dot{{background:var(--phys-mid);}}
[data-tab="double"] .tab-dot{{background:var(--double-mid);}}
[data-tab="sa"] .tab-dot{{background:var(--sa-mid);}}
.filter-bar{{background:white;border-bottom:1px solid var(--border);padding:10px 24px;}}
.filter-inner{{max-width:1200px;margin:0 auto;display:flex;align-items:center;gap:10px;flex-wrap:wrap;}}
.filter-label{{font-size:11px;font-weight:700;color:var(--text-light);text-transform:uppercase;letter-spacing:.8px;}}
.filter-divider{{width:1px;height:20px;background:var(--border);margin:0 4px;flex-shrink:0;}}
.filter-btn{{padding:5px 14px;border-radius:20px;border:1px solid var(--border);background:white;font-size:13px;cursor:pointer;transition:all .15s;color:var(--text-mid);font-weight:500;}}
.filter-btn:hover{{border-color:#999;background:#fafafa;}}
.filter-btn.active{{background:var(--text-dark);color:white;border-color:var(--text-dark);}}
.filter-btn.subject-bio.active{{background:var(--bio-color);border-color:var(--bio-color);color:white;}}
.filter-btn.subject-chem.active{{background:var(--chem-color);border-color:var(--chem-color);color:white;}}
.filter-btn.subject-phys.active{{background:var(--phys-color);border-color:var(--phys-color);color:white;}}
.subject-filters{{display:none;align-items:center;gap:8px;}}
.subject-filters.visible{{display:flex;}}
.search-box{{margin-left:auto;padding:7px 14px 7px 36px;border:1px solid var(--border);border-radius:20px;font-size:13px;outline:none;width:240px;transition:border-color .2s,box-shadow .2s;background:#fafafa url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' width='14' height='14' viewBox='0 0 24 24' fill='none' stroke='%23999' stroke-width='2.5'%3E%3Ccircle cx='11' cy='11' r='8'/%3E%3Cpath d='m21 21-4.35-4.35'/%3E%3C/svg%3E") no-repeat 13px center;}}
.search-box:focus{{border-color:#666;box-shadow:0 0 0 2px rgba(0,0,0,.08);background-color:white;}}
.content-area{{max-width:1200px;margin:0 auto;padding:28px 16px;}}
.section-panel{{display:none;}}
.section-panel.active{{display:block;}}
.section-intro{{display:flex;align-items:flex-start;gap:14px;margin-bottom:20px;padding:16px 20px;border-radius:12px;font-size:14px;}}
.section-intro-biology{{background:var(--bio-light);color:var(--bio-color);border-left:4px solid var(--bio-mid);}}
.section-intro-chemistry{{background:var(--chem-light);color:var(--chem-color);border-left:4px solid var(--chem-mid);}}
.section-intro-physics{{background:var(--phys-light);color:var(--phys-color);border-left:4px solid var(--phys-mid);}}
.section-intro-double{{background:var(--double-light);color:var(--double-color);border-left:4px solid var(--double-mid);}}
.section-intro-sa{{background:var(--sa-light);color:var(--sa-color);border-left:4px solid var(--sa-mid);}}
.intro-icon{{font-size:24px;flex-shrink:0;margin-top:1px;}}
.intro-text strong{{display:block;font-size:15px;font-weight:700;margin-bottom:3px;}}
.stats-row{{display:flex;gap:10px;margin-bottom:18px;flex-wrap:wrap;}}
.stat-chip{{background:white;border:1px solid var(--border);border-radius:8px;padding:8px 16px;font-size:13px;color:var(--text-mid);display:flex;align-items:center;gap:6px;}}
.stat-num{{font-weight:700;font-size:15px;color:var(--text-dark);}}
.cards-grid{{display:grid;grid-template-columns:repeat(auto-fill,minmax(310px,1fr));gap:14px;grid-auto-rows:780px;max-width:1200px;}}
.lesson-card{{background:var(--card-bg);border-radius:12px;box-shadow:var(--shadow);overflow:hidden;border:1px solid var(--border);transition:box-shadow .2s,transform .15s;display:flex;flex-direction:column;height:100%;}}
.lesson-card:hover{{box-shadow:0 6px 20px rgba(0,0,0,.12);transform:translateY(-1px);}}
.lesson-card.hidden{{display:none;}}
.card-header{{padding:12px 14px;min-height:80px;display:flex;flex-direction:column;justify-content:center;}}
.card-header-biology{{background:linear-gradient(135deg,#1b5e35 0%,#2d6a4f 100%);}}
.card-header-chemistry{{background:linear-gradient(135deg,#7c2f00 0%,#b84400 100%);}}
.card-header-physics{{background:linear-gradient(135deg,#0d3d6e 0%,#1565c0 100%);}}
.card-title-row{{display:flex;justify-content:space-between;align-items:flex-start;gap:8px;}}
.lesson-link{{color:white;text-decoration:none;display:flex;align-items:flex-start;gap:9px;flex:1;}}
.lesson-link:hover .lesson-name{{text-decoration:underline;text-underline-offset:2px;}}
.lesson-label{{font-size:11px;font-weight:700;background:rgba(255,255,255,.2);padding:3px 8px;border-radius:10px;white-space:nowrap;flex-shrink:0;letter-spacing:.3px;margin-top:1px;}}
.lesson-name{{font-size:13px;font-weight:600;color:white;flex:1;line-height:1.35;}}
.link-arrow{{font-size:14px;color:rgba(255,255,255,.6);flex-shrink:0;transition:transform .15s;margin-top:1px;}}
.lesson-link:hover .link-arrow{{transform:translateX(3px);color:white;}}
.year-badge{{font-size:10px;background:rgba(0,0,0,.2);color:rgba(255,255,255,.88);padding:2px 8px;border-radius:10px;white-space:nowrap;flex-shrink:0;font-weight:500;margin-top:2px;}}
.card-body{{padding:10px 13px;display:flex;flex-direction:column;position:relative;flex:1;min-height:0;overflow-y:auto;}}
.card-body::-webkit-scrollbar{{width:5px;}}
.card-body::-webkit-scrollbar-thumb{{background:#ccc;border-radius:4px;}}
.card-body::-webkit-scrollbar-thumb:hover{{background:#aaa;}}
.card-top{{position:relative;}}
.student-folder-btn{{display:flex;align-items:center;justify-content:center;gap:6px;margin:0 0 10px;padding:8px 14px;border-radius:8px;background:#e8f0fe;color:#1565c0;border:1px solid #bbdefb;font-size:12px;font-weight:600;text-decoration:none;transition:background .15s;}}
.student-folder-btn:hover{{background:#bbdefb;color:#0d47a1;}}
.info-scroll{{max-height:600px;overflow-y:auto;padding-right:4px;}}
.info-scroll::-webkit-scrollbar{{width:4px;}}
.info-scroll::-webkit-scrollbar-thumb{{background:#ccc;border-radius:4px;}}
.student-section{{margin:0 0 10px;padding:8px 12px;border-radius:8px;font-size:12px;line-height:1.5;}}
.student-section:last-of-type{{margin-bottom:14px;}}
.student-section-title{{font-weight:700;font-size:12px;margin-bottom:4px;}}
.student-section ul{{margin:0;padding:0 0 0 18px;}}
.student-section li{{margin-bottom:3px;}}
.student-objectives{{background:#e8f5e9;border:1px solid #a5d6a7;color:#1b5e20;}}
.student-objectives .student-section-title{{color:#2e7d32;}}
.student-keywords{{background:#fff3e0;border:1px solid #ffcc80;color:#e65100;}}
.student-keywords .student-section-title{{color:#ef6c00;}}
.student-specs{{background:#e3f2fd;border:1px solid #90caf9;color:#0d47a1;}}
.student-specs .student-section-title{{color:#1565c0;}}
.student-specs .spec-list{{list-style:none;padding:0;display:flex;flex-direction:column;gap:6px;}}
.student-specs .spec-item{{display:flex;gap:8px;align-items:flex-start;font-size:12px;line-height:1.45;}}
.student-specs .spec-num{{font-weight:700;font-size:10px;background:rgba(255,255,255,.7);color:#0d47a1;padding:2px 6px;border-radius:5px;white-space:nowrap;flex-shrink:0;font-family:'Courier New',monospace;margin-top:1px;border:1px solid #bbdefb;}}
.student-specs .spec-text{{color:#1a237e;flex:1;}}
.student-safety{{background:#fce4ec;border:1px solid #ef9a9a;color:#b71c1c;}}
.student-safety .student-section-title{{color:#c62828;}}
.student-safety-box{{margin:10px 0 4px;padding:10px 14px;border-radius:8px;background:#fbe9e7;border:1px solid #ef9a9a;font-size:12px;line-height:1.5;color:#b71c1c;}}
.student-safety-box strong{{margin-left:4px;font-size:12px;}}
.safety-line{{margin-top:4px;padding-left:22px;color:#4e342e;}}
.card-pages{{display:flex;align-items:center;gap:7px;font-size:12px;color:var(--text-light);font-weight:600;margin-bottom:12px;padding-bottom:12px;border-bottom:1px solid var(--border);}}
.spec-list{{list-style:none;padding:0;display:flex;flex-direction:column;gap:9px;}}
.spec-item{{display:flex;gap:8px;align-items:flex-start;font-size:13px;line-height:1.55;}}
.spec-num{{font-weight:700;font-size:11px;background:#eef0f5;color:#333;padding:2px 8px;border-radius:6px;white-space:nowrap;flex-shrink:0;font-family:'Courier New',monospace;margin-top:2px;border:1px solid #dde0e8;}}
.spec-text{{color:var(--text-mid);flex:1;}}
.badge{{display:inline-flex;align-items:center;font-size:10px;font-weight:700;padding:1px 7px;border-radius:4px;margin-left:5px;white-space:nowrap;flex-shrink:0;margin-top:2px;}}
.badge-bio{{background:var(--bio-light);color:var(--bio-color);}}
.badge-chem{{background:var(--chem-light);color:var(--chem-color);}}
.badge-phys{{background:var(--phys-light);color:var(--phys-color);}}
.no-results{{text-align:center;padding:56px 16px;color:var(--text-light);}}
.subject-divider{{display:flex;align-items:center;gap:12px;padding:14px 4px 6px;margin-top:4px;font-size:14px;font-weight:700;}}
.subject-divider::after{{content:'';flex:1;height:2px;border-radius:1px;}}
.subject-divider-bio{{color:var(--bio-color);}}.subject-divider-bio::after{{background:var(--bio-light);}}
.subject-divider-chem{{color:var(--chem-color);}}.subject-divider-chem::after{{background:var(--chem-light);}}
.subject-divider-phys{{color:var(--phys-color);}}.subject-divider-phys::after{{background:var(--phys-light);}}
.jump-nav{{display:flex;align-items:center;gap:8px;margin-bottom:12px;flex-wrap:wrap;}}
.jump-label{{font-size:11px;font-weight:700;color:var(--text-light);text-transform:uppercase;letter-spacing:.8px;margin-right:2px;}}
.jump-btn{{padding:5px 14px;border-radius:20px;border:1px solid;font-size:13px;font-weight:600;text-decoration:none;cursor:pointer;transition:all .15s;}}
.jump-btn-bio{{color:var(--bio-color);border-color:var(--bio-mid);background:var(--bio-light);}}.jump-btn-bio:hover{{background:var(--bio-mid);color:white;}}
.jump-btn-chem{{color:var(--chem-color);border-color:var(--chem-mid);background:var(--chem-light);}}.jump-btn-chem:hover{{background:var(--chem-mid);color:white;}}
.jump-btn-phys{{color:var(--phys-color);border-color:var(--phys-mid);background:var(--phys-light);}}.jump-btn-phys:hover{{background:var(--phys-mid);color:white;}}
.no-results-icon{{font-size:40px;margin-bottom:12px;}}
.no-results p{{font-size:15px;}}
.site-footer{{text-align:center;padding:20px;font-size:11px;color:var(--text-light);border-top:1px solid var(--border);margin-top:20px;}}
.assess-dropdown-wrap{{position:relative;display:inline-block;margin-top:10px;}}
.assess-btn{{cursor:pointer;background:rgba(255,255,255,.1) !important;border:1px solid rgba(255,255,255,.3) !important;color:rgba(255,255,255,.85) !important;font-size:12px !important;font-weight:700;padding:6px 16px !important;border-radius:20px !important;transition:background .2s;}}
.assess-btn:hover{{background:rgba(255,255,255,.2) !important;}}
.assess-panel{{display:none;position:absolute;top:100%;left:50%;transform:translateX(-50%);margin-top:8px;background:white;border-radius:12px;box-shadow:0 10px 40px rgba(0,0,0,.25);padding:16px;z-index:200;min-width:340px;max-width:500px;max-height:70vh;overflow-y:auto;}}
.assess-panel.open{{display:block;}}
.assess-row.assess-next-y10{{background:#e3f2fd;font-weight:700;}}
.assess-row.assess-next-y10 td{{color:#1565c0;}}
.assess-row.assess-next-y11{{background:#fff3e0;font-weight:700;}}
.assess-row.assess-next-y11 td{{color:#e65100;}}
.assess-next-marker{{display:inline-block;font-size:10px;margin-right:4px;}}
.assess-next-y10 .assess-next-marker{{color:#1e88e5;}}
.assess-next-y11 .assess-next-marker{{color:#fb8c00;}}
.assess-yg-label{{display:inline-block;font-size:9px;font-weight:700;padding:1px 5px;border-radius:8px;margin-left:6px;vertical-align:middle;}}
.assess-next-y10 .assess-yg-label{{background:#bbdefb;color:#1565c0;}}
.assess-next-y11 .assess-yg-label{{background:#ffe0b2;color:#e65100;}}
.assess-subject{{margin-bottom:12px;}}
.assess-subject:last-child{{margin-bottom:0;}}
.assess-subject-title{{font-weight:700;font-size:13px;margin-bottom:6px;padding-bottom:4px;border-bottom:2px solid #eee;}}
.assess-title-bio{{color:var(--bio-color);border-color:var(--bio-light);}}
.assess-title-chem{{color:var(--chem-color);border-color:var(--chem-light);}}
.assess-title-phys{{color:var(--phys-color);border-color:var(--phys-light);}}
.assess-table{{width:100%;border-collapse:collapse;font-size:12px;}}
.assess-table th{{text-align:left;padding:4px 8px;font-size:10px;text-transform:uppercase;letter-spacing:.5px;color:var(--text-light);border-bottom:1px solid #eee;}}
.assess-table td{{padding:5px 8px;color:var(--text-dark);border-bottom:1px solid #f5f5f5;}}
.assess-table tr:last-child td{{border-bottom:none;}}
.assess-row:hover{{background:#f8f9fa;}}
.assess-tabs{{display:flex;gap:2px;margin-bottom:12px;border-bottom:2px solid #eee;padding-bottom:0;overflow-x:auto;}}
.assess-tab{{background:none;border:none;padding:6px 12px;font-size:11px;font-weight:600;color:var(--text-light);cursor:pointer;border-bottom:2px solid transparent;margin-bottom:-2px;white-space:nowrap;transition:color .2s,border-color .2s;}}
.assess-tab:hover{{color:var(--text-dark);}}
.assess-tab-active{{color:var(--text-dark);border-bottom-color:var(--bio-color);}}
.assess-empty{{color:var(--text-light);font-size:12px;font-style:italic;padding:12px 0;text-align:center;}}
.assess-ds-subject{{margin-bottom:10px;}}
.assess-ds-subject:last-child{{margin-bottom:0;}}
.assess-ds-subject .assess-subject-title{{font-size:12px;margin-bottom:4px;padding-bottom:3px;border-bottom:2px solid #eee;}}
.subj-dot{{display:inline-block;width:8px;height:8px;border-radius:50%;margin-right:5px;}}
.subj-dot-bio{{background:var(--bio-mid);}}
.subj-dot-chem{{background:var(--chem-mid);}}
.subj-dot-phys{{background:var(--phys-mid);}}
@media(max-width:768px){{.site-header{{padding:16px 14px 12px;}}.header-title{{font-size:20px !important;margin-bottom:4px;}}.header-subtitle{{font-size:11px;margin-bottom:0;}}.header-eyebrow{{padding:3px 10px;font-size:9px;margin-bottom:6px;}}.spec-codes{{gap:5px;margin-top:10px;}}.spec-tag{{padding:3px 8px;font-size:10px;}}.nav-wrapper{{position:static;}}.nav-tabs{{flex-wrap:wrap;gap:2px;padding:0 8px;}}.nav-tab{{padding:8px 12px;font-size:12px;}}.filter-bar{{padding:8px 10px;}}.filter-inner{{flex-direction:row;flex-wrap:wrap;gap:6px;align-items:center;}}.filter-btn{{padding:4px 10px;font-size:11px;}}.search-box{{width:100%;margin-left:0;margin-top:6px;}}.content-area{{padding:10px 8px;}}.section-intro{{padding:10px 14px;margin-bottom:10px;font-size:12px;gap:10px;}}.intro-icon{{font-size:18px;}}.intro-text strong{{font-size:13px;}}.stats-row{{margin-bottom:10px;gap:6px;}}.stat-chip{{padding:4px 10px;font-size:11px;}}.cards-grid{{grid-template-columns:1fr !important;grid-auto-rows:auto !important;gap:10px;}}.lesson-card{{height:auto !important;}}.info-scroll{{max-height:260px;overflow-y:auto;}}.card-header{{padding:10px 12px;min-height:auto;}}.lesson-name{{font-size:12px;}}.student-section{{padding:6px 10px;margin:0 0 6px;}}.student-section-title{{font-size:11px;}}.student-section li,.student-section p{{font-size:11px;}}.student-folder-btn{{padding:6px 10px;font-size:11px;margin:0 0 8px;}}.card-body{{padding:8px 10px;}}.card-pages{{font-size:11px;margin-bottom:8px;padding-bottom:8px;}}}}

/* ── Print stylesheet ── */
@media print {{
  .nav-wrapper, .filter-bar, .card-modal-overlay, .search-box,
  .edition-link, .teacher-banner, .copy-btn, .student-folder-btn,
  .tsec-arrow, .section-intro, .stats-row, .site-footer {{ display:none !important; }}
  body {{ background:white; color:black; }}
  .site-header {{ background:#1a1a2e !important; padding:16px 20px 12px; -webkit-print-color-adjust:exact; print-color-adjust:exact; }}
  .site-header::before {{ display:none; }}
  .header-title {{ font-size:20px !important; margin-bottom:4px; }}
  .header-subtitle, .spec-codes {{ display:none; }}
  .header-eyebrow {{ font-size:10px; }}
  .content-area {{ padding:0; }}
  .cards-grid {{ display:grid; grid-template-columns:repeat(2,1fr) !important; grid-auto-rows:auto !important; gap:8px; max-width:100%; }}
  .lesson-card {{ height:auto !important; break-inside:avoid; page-break-inside:avoid; border:1px solid #ccc; box-shadow:none; }}
  .card-header {{ min-height:auto; padding:8px 10px; -webkit-print-color-adjust:exact; print-color-adjust:exact; }}
  .lesson-name {{ font-size:11px; }}
  .card-body {{ padding:6px 8px; }}
  .info-scroll {{ max-height:none !important; overflow:visible !important; }}
  .student-section {{ padding:4px 8px; margin:0 0 4px; font-size:10px; -webkit-print-color-adjust:exact; print-color-adjust:exact; }}
  .student-section-title {{ font-size:10px; }}
  .student-section li, .student-section p {{ font-size:10px; }}
  .card-pages {{ font-size:10px; }}
  .card-bottom {{ max-height:none; overflow:visible; }}
  .teacher-section {{ break-inside:avoid; }}
  .tsec-body {{ max-height:none !important; overflow:visible !important; }}
  .tsec-body.collapsed {{ display:block !important; max-height:none !important; }}
  .tsec-header {{ padding:6px 10px; font-size:11px; -webkit-print-color-adjust:exact; print-color-adjust:exact; }}
  .subject-divider {{ break-before:page; }}
  .tab-panel:not(.active) {{ display:none; }}
}}

/* ── Full-screen card modal ── */
.card-modal-overlay{{position:fixed;inset:0;background:rgba(0,0,0,.55);z-index:1000;display:flex;align-items:center;justify-content:center;padding:20px;opacity:0;pointer-events:none;transition:opacity .2s ease;}}
.card-modal-overlay.open{{opacity:1;pointer-events:all;}}
.card-modal{{background:var(--card-bg);border-radius:14px;box-shadow:0 20px 60px rgba(0,0,0,.35);width:100%;max-width:820px;max-height:90vh;display:flex;flex-direction:column;overflow:hidden;transform:scale(.96);transition:transform .2s ease;}}
.card-modal-overlay.open .card-modal{{transform:scale(1);}}
.card-modal-header{{padding:14px 16px;flex-shrink:0;position:relative;}}
.card-modal-close{{position:absolute;top:14px;right:16px;background:rgba(255,255,255,.2);border:none;color:white;border-radius:50%;width:28px;height:28px;font-size:16px;cursor:pointer;display:flex;align-items:center;justify-content:center;line-height:1;transition:background .15s;}}
.card-modal-close:hover{{background:rgba(255,255,255,.4);}}
.card-modal-body{{padding:14px 16px;overflow-y:auto;flex:1;}}
.card-modal-body .info-scroll{{max-height:none;overflow:visible;}}
.card-modal-body .student-folder-btn{{display:flex;}}
.card-modal-cursor{{cursor:pointer;}}
.card-modal-cursor:hover{{box-shadow:0 8px 24px rgba(0,0,0,.15);transform:translateY(-2px);}}
{teacher_extra}
"""

JS = """
const tabs = document.querySelectorAll('.nav-tab');
const panels = document.querySelectorAll('.section-panel');
const subjectFiltersRow = document.getElementById('subject-filters-row');
tabs.forEach(tab => {
  tab.addEventListener('click', () => {
    tabs.forEach(t => t.classList.remove('active'));
    panels.forEach(p => p.classList.remove('active'));
    tab.classList.add('active');
    document.getElementById('panel-' + tab.dataset.tab).classList.add('active');
    subjectFiltersRow.classList.toggle('visible', tab.dataset.tab === 'double' || tab.dataset.tab === 'sa');
    document.querySelectorAll('.filter-btn[data-subject]').forEach(b => b.classList.remove('active'));
    document.querySelector('.filter-btn[data-subject="all"]').classList.add('active');
    activeSubject = 'all';
    document.querySelectorAll('.filter-btn[data-term]').forEach(b => b.classList.remove('active'));
    document.querySelector('.filter-btn[data-term="all"]').classList.add('active');
    activeTerm = 'all';
    applyFilters();
  });
});
const yearBtns = document.querySelectorAll('.filter-btn[data-year]');
let activeYear = 'all';
yearBtns.forEach(btn => {
  btn.addEventListener('click', () => {
    yearBtns.forEach(b => b.classList.remove('active'));
    btn.classList.add('active');
    activeYear = btn.dataset.year;
    applyFilters();
  });
});
const termBtns = document.querySelectorAll('.filter-btn[data-term]');
let activeTerm = 'all';
termBtns.forEach(btn => {
  btn.addEventListener('click', () => {
    termBtns.forEach(b => b.classList.remove('active'));
    btn.classList.add('active');
    activeTerm = btn.dataset.term;
    applyFilters();
  });
});
const subjectBtns = document.querySelectorAll('.filter-btn[data-subject]');
let activeSubject = 'all';
subjectBtns.forEach(btn => {
  btn.addEventListener('click', () => {
    subjectBtns.forEach(b => b.classList.remove('active'));
    btn.classList.add('active');
    activeSubject = btn.dataset.subject;
    applyFilters();
    /* In Double/SA panels, scroll to the subject anchor */
    const activeTab = document.querySelector('.nav-tab.active');
    if (activeTab && (activeTab.dataset.tab === 'double' || activeTab.dataset.tab === 'sa')) {
      const subject = btn.dataset.subject;
      if (subject !== 'all') {
        const anchor = document.getElementById('jump-' + activeTab.dataset.tab + '-' + subject);
        if (anchor) setTimeout(() => anchor.scrollIntoView({behavior:'smooth', block:'start'}), 50);
      }
    }
  });
});
const searchBox = document.getElementById('searchBox');
let searchTimer;
searchBox.addEventListener('input', () => { clearTimeout(searchTimer); searchTimer = setTimeout(applyFilters, 300); });
function applyFilters() {
  const query = searchBox.value.toLowerCase().trim();
  ['biology','chemistry','physics','double','sa'].forEach(sec => {
    const grid = document.getElementById('grid-' + sec);
    const noResults = document.getElementById('noresults-' + sec);
    if (!grid) return;
    const cards = grid.querySelectorAll('.lesson-card');
    let visible = 0;
    const needsSub = (sec === 'double' || sec === 'sa');
    cards.forEach(card => {
      const ok = (activeYear === 'all' || card.dataset.year === activeYear)
              && (activeTerm === 'all' || card.dataset.term === activeTerm)
              && (!needsSub || activeSubject === 'all' || card.dataset.subject === activeSubject)
              && (!query || card.textContent.toLowerCase().includes(query));
      card.classList.toggle('hidden', !ok);
      if (ok) visible++;
    });
    if (noResults) noResults.classList.toggle('hidden', visible > 0);
    updateStats(sec, visible, cards.length);
  });
}
function updateStats(sec, visible, total) {
  const el = document.getElementById('stats-' + sec);
  if (!el) return;
  const label = (visible === total) ? (total + ' lessons') : (visible + ' of ' + total + ' lessons');
  let extra = '';
  if (sec === 'double' || sec === 'sa') {
    const vis = [...document.getElementById('grid-'+sec).querySelectorAll('.lesson-card:not(.hidden)')];
    extra = '<div class="stat-chip"><span class="subj-dot subj-dot-bio"></span><span class="stat-num" style="color:var(--bio-color)">'
          + vis.filter(c=>c.dataset.subject==='biology').length + '</span>&nbsp;Biology</div>'
          + '<div class="stat-chip"><span class="subj-dot subj-dot-chem"></span><span class="stat-num" style="color:var(--chem-color)">'
          + vis.filter(c=>c.dataset.subject==='chemistry').length + '</span>&nbsp;Chemistry</div>'
          + '<div class="stat-chip"><span class="subj-dot subj-dot-phys"></span><span class="stat-num" style="color:var(--phys-color)">'
          + vis.filter(c=>c.dataset.subject==='physics').length + '</span>&nbsp;Physics</div>';
  }
  el.innerHTML = '<div class="stat-chip"><span class="stat-num">' + label + '</span></div>' + extra;
}
applyFilters();

/* ── Card expand (student view) ── */
function initCards() {
  document.querySelectorAll('.lesson-card').forEach(card => {
    const body = card.querySelector('.card-body');
    const btn  = card.querySelector('.card-expand-btn');
    const fade = card.querySelector('.card-fade');
    if (!btn || !body) return;
    /* Show expand button whenever content is taller than the compact max-height */
    if (body.scrollHeight > body.clientHeight + 2) {
      btn.classList.add('shown');
      if (fade) fade.classList.add('show');
    }
  });
}
function toggleExpand(btn) {
  const card     = btn.closest('.lesson-card');
  const fade     = card.querySelector('.card-fade');
  const expanded = card.classList.toggle('card-expanded');
  btn.innerHTML  = expanded ? '&#9650; Show less' : '&#9660; Show more';
  if (fade) fade.classList.toggle('show', !expanded);
}
/* Run after layout is painted */
requestAnimationFrame(() => requestAnimationFrame(initCards));

/* ── Accordion toggle ── */
function toggleSection(header) {
  const body  = header.nextElementSibling;
  const isOpen = !body.classList.contains('collapsed');
  body.classList.toggle('collapsed', isOpen);
  header.classList.toggle('open', !isOpen);
}

/* ── Copy button ── */
function copySection(event, btn) {
  event.stopPropagation();
  const plainText  = btn.dataset.plain;
  const htmlContent = btn.dataset.copyhtml || null;

  function showSuccess() {
    const orig = btn.innerHTML;
    btn.innerHTML = '&#10003; Copied!';
    btn.classList.add('copied');
    setTimeout(() => { btn.innerHTML = orig; btn.classList.remove('copied'); }, 2000);
  }

  /* Rich copy (plain + HTML) — links stay clickable when pasted into Word/Outlook/Gmail */
  if (htmlContent && navigator.clipboard && window.ClipboardItem) {
    navigator.clipboard.write([
      new ClipboardItem({
        'text/plain': new Blob([plainText],   {type: 'text/plain'}),
        'text/html':  new Blob([htmlContent], {type: 'text/html'})
      })
    ]).then(showSuccess).catch(() => {
      /* ClipboardItem blocked (e.g. non-HTTPS) — fall back to plain text */
      navigator.clipboard.writeText(plainText).then(showSuccess);
    });
  } else {
    /* Plain-text copy (Requisitions, or browsers without ClipboardItem) */
    navigator.clipboard.writeText(plainText).then(showSuccess).catch(() => {
      const ta = document.createElement('textarea');
      ta.value = plainText;
      document.body.appendChild(ta);
      ta.select();
      document.execCommand('copy');
      document.body.removeChild(ta);
      showSuccess();
    });
  }
}
/* ── Full-screen card modal ── */
function openCardModal(card) {
  const header = card.querySelector('.card-header');
  const body   = card.querySelector('.card-body');
  if (!header || !body) return;

  // Clone header and body into modal
  const mHeader = document.getElementById('cardModalHeader');
  const mBody   = document.getElementById('cardModalBody');
  mHeader.innerHTML = header.innerHTML;
  mHeader.className = 'card-modal-header ' + header.className.replace('card-header','');
  mBody.innerHTML   = body.innerHTML;

  // Re-enable any collapsed accordions so content is visible
  mBody.querySelectorAll('.tsec-body.collapsed').forEach(b => {
    b.classList.remove('collapsed');
    const h = b.previousElementSibling;
    if (h) h.classList.add('open');
  });

  document.getElementById('cardModal').classList.add('open');
  document.getElementById('cardModalBody').scrollTop = 0;
  document.body.style.overflow = 'hidden';
}
function closeCardModal(e) {
  const inner = document.getElementById('cardModalInner');
  if (!inner.contains(e.target)) {
    document.getElementById('cardModal').classList.remove('open');
    document.body.style.overflow = '';
  }
}
document.addEventListener('keydown', e => {
  if (e.key === 'Escape') {
    document.getElementById('cardModal').classList.remove('open');
    document.body.style.overflow = '';
    const ap = document.getElementById('assessPanel');
    if (ap) ap.classList.remove('open');
  }
});
/* Switch assessment date tabs */
function switchAssessTab(btn) {
  const key = btn.dataset.assess;
  btn.closest('.assess-panel').querySelectorAll('.assess-tab').forEach(t => t.classList.remove('assess-tab-active'));
  btn.classList.add('assess-tab-active');
  btn.closest('.assess-panel').querySelectorAll('.assess-tab-panel').forEach(p => p.style.display = 'none');
  document.getElementById('assess-' + key).style.display = '';
}
/* Highlight next upcoming test per year group (Y10=blue tests 1-4, Y11=orange tests 5+) */
(function() {
  const today = new Date().toISOString().slice(0,10);
  document.querySelectorAll('.assess-table').forEach(table => {
    [['y10','Year 10'],['y11','Year 11']].forEach(([yg, label]) => {
      const cls = 'assess-next-' + yg;
      const rows = [...table.querySelectorAll(`.assess-row[data-yg="${yg}"]`)]
        .sort((a,b) => (a.dataset.start||'').localeCompare(b.dataset.start||''));
      for (const row of rows) {
        const end = row.dataset.end || row.dataset.start;
        if (end && end >= today) {
          row.classList.add(cls);
          const firstTd = row.querySelector('td');
          if (firstTd) firstTd.innerHTML =
            '<span class="assess-next-marker">&#9654;</span>' +
            firstTd.textContent +
            '<span class="assess-yg-label">' + label + '</span>';
          break;
        }
      }
    });
  });
})();
/* Close assessment dropdown when clicking outside */
document.addEventListener('click', e => {
  const ap = document.getElementById('assessPanel');
  if (ap && ap.classList.contains('open') && !e.target.closest('.assess-dropdown-wrap')) {
    ap.classList.remove('open');
  }
});
// Attach click listener to all cards
document.querySelectorAll('.lesson-card').forEach(card => {
  card.classList.add('card-modal-cursor');
  card.addEventListener('click', function(e) {
    // Don't open modal if clicking a link, button, or accordion header
    if (e.target.closest('a, button, .tsec-header, .copy-btn, .filter-btn')) return;
    openCardModal(this);
  });
});

"""

SUBJ_FILTER_HTML = """
    <div class="subject-filters" id="subject-filters-row">
      <span class="filter-label">Subject</span>
      <button class="filter-btn active" data-subject="all">All</button>
      <button class="filter-btn subject-bio"  data-subject="biology">&#127807; Biology</button>
      <button class="filter-btn subject-chem" data-subject="chemistry">&#9879;&#65039; Chemistry</button>
      <button class="filter-btn subject-phys" data-subject="physics">&#128302; Physics</button>
    </div>"""

# ── Panel builder ──────────────────────────────────────────────────────────────
def section_panel(panel_id, intro_class, icon, title, desc, stats_id, grid_id, cards_html, noresults_id, active=False, extra_html='', grid_wrapper=True):
    active_class = ' active' if active else ''
    # For single-subject panels use a cards-grid wrapper; for multi-subject use a plain div
    open_tag  = f'<div class="cards-grid" id="{grid_id}">' if grid_wrapper else f'<div id="{grid_id}">'
    return f"""
  <div class="section-panel{active_class}" id="panel-{panel_id}">
    <div class="section-intro section-intro-{intro_class}">
      <span class="intro-icon">{icon}</span>
      <div class="intro-text"><strong>{title}</strong>{desc}</div>
    </div>
    <div class="stats-row" id="{stats_id}"></div>
    {open_tag}{cards_html}</div>
    <div class="no-results hidden" id="{noresults_id}">
      <div class="no-results-icon">&#128269;</div><p>No more lessons match your current filters.</p>
    </div>
  </div>"""

# ── HTML builder ───────────────────────────────────────────────────────────────
def build_html(data, sa_sets, teacher=False):
    timestamp = datetime.now().strftime('%d %b %Y %H:%M')

    # Build assessment dates dropdown with tabbed sub-sections
    assessments = data.get('_assessments', {})

    def _build_table(tests):
        if not tests:
            return '<p class="assess-empty">No dates set yet</p>'
        rows = ''.join(
            f'<tr class="assess-row" data-yg="{t.get("year_group","y10")}" data-start="{t.get("start_iso","")}" data-end="{t.get("end_iso","")}">'
            f'<td>{t["name"]}</td><td>{t["start"]}</td><td>{t["end"]}</td></tr>'
            for t in tests
        )
        return (f'<table class="assess-table"><thead><tr><th>Test</th><th>Start</th><th>End</th></tr></thead>'
                f'<tbody>{rows}</tbody></table>')

    # Build the Double & Single Award panel with sub-tables for each subject
    def _build_ds_panel():
        ds_bio = assessments.get('DS_Biology', [])
        ds_chem = assessments.get('DS_Chemistry', [])
        ds_phys = assessments.get('DS_Physics', [])
        if not ds_bio and not ds_chem and not ds_phys:
            return '<p class="assess-empty">No dates set yet</p>'
        html = ''
        for label, tests, cls in [('&#127807; Biology', ds_bio, 'bio'), ('&#9879;&#65039; Chemistry', ds_chem, 'chem'), ('&#128302; Physics', ds_phys, 'phys')]:
            if tests:
                html += f'<div class="assess-ds-subject"><div class="assess-subject-title assess-title-{cls}">{label}</div>{_build_table(tests)}</div>'
        return html

    # Define the 4 tabs: Bio, Chem, Phys (single subject) + Double & Single Award (combined)
    assess_tabs_data = [
        ('bio', '&#127807; Biology', _build_table(assessments.get('Biology', []))),
        ('chem', '&#9879;&#65039; Chemistry', _build_table(assessments.get('Chemistry', []))),
        ('phys', '&#128302; Physics', _build_table(assessments.get('Physics', []))),
        ('dsaward', '&#128218; Double &amp; Single Award', _build_ds_panel()),
    ]

    # Build tab buttons and panels
    tab_btns = ''
    tab_panels = ''
    for i, (key, label, content) in enumerate(assess_tabs_data):
        active = ' assess-tab-active' if i == 0 else ''
        hidden = '' if i == 0 else ' style="display:none"'
        tab_btns += f'<button class="assess-tab{active}" data-assess="{key}" onclick="switchAssessTab(this)">{label}</button>'
        tab_panels += f'<div class="assess-tab-panel" id="assess-{key}"{hidden}>{content}</div>'

    assess_dropdown = (
        '<div class="assess-dropdown-wrap">'
        '<button class="spec-tag assess-btn" onclick="document.getElementById(\'assessPanel\').classList.toggle(\'open\')">'
        '&#128197; Assessment Dates &#9662;</button>'
        '<div class="assess-panel" id="assessPanel">'
        f'<div class="assess-tabs">{tab_btns}</div>'
        f'{tab_panels}'
        '</div></div>'
    )
    mode_label = 'TEACHER EDITION' if teacher else ''
    teacher_note = ' &mdash; includes requisitions &amp; teacher notes' if teacher else ''

    # Build cards
    single_cards = {}
    dbl_by_subject = {s: [] for s in ['Biology', 'Chemistry', 'Physics']}
    sa_by_subject  = {s: [] for s in ['Biology', 'Chemistry', 'Physics']}

    SUBJ_ICONS   = {'Biology': '&#127807;', 'Chemistry': '&#9879;&#65039;', 'Physics': '&#128302;'}
    SUBJ_DIV_CLS = {'Biology': 'bio', 'Chemistry': 'chem', 'Physics': 'phys'}

    for subject in ['Biology', 'Chemistry', 'Physics']:
        s_cards = []
        for l in data[subject]:
            s_cards.append(render_card(l, subject, teacher=teacher))
        single_cards[subject] = '\n'.join(s_cards)

        for l in data[subject]:
            sp = filter_double(l)
            if sp is not None:
                dbl_by_subject[subject].append(render_card(l, subject, sp, teacher=teacher))

        for l in data[subject]:
            sp = filter_sa(l, subject, sa_sets)
            if sp is not None:
                sa_by_subject[subject].append(render_card(l, subject, sp, teacher=teacher))

    def grouped_html(by_subject, panel_id):
        """Each subject gets its own divider heading + its own mini cards-grid (no gap issues)."""
        out = ''
        for subj in ['Biology', 'Chemistry', 'Physics']:
            sl  = subj.lower()
            sc  = SUBJ_DIV_CLS[subj]
            ico = SUBJ_ICONS[subj]
            cards = '\n'.join(by_subject[subj])
            out += (f'<div class="subject-divider subject-divider-{sc}" id="jump-{panel_id}-{sl}">'
                    f'<span>{ico} {subj}</span></div>\n'
                    f'<div class="cards-grid">{cards}</div>\n')
        return out

    double_html = grouped_html(dbl_by_subject, 'double')
    sa_html     = grouped_html(sa_by_subject,  'sa')

    panels = ''
    tab_subjects = [
        ('biology',   'biology',   '&#127807;',  'Single Science Biology &mdash; 4BI0',
         'All Biology lessons including Biology-only spec points.',
         single_cards['Biology'], True),
        ('chemistry', 'chemistry', '&#9879;&#65039;', 'Single Science Chemistry &mdash; 4CH0',
         'All Chemistry lessons including Chemistry-only spec points.',
         single_cards['Chemistry'], False),
        ('physics',   'physics',   '&#128302;',  'Single Science Physics &mdash; 4PH0',
         'All Physics lessons including Physics-only spec points.',
         single_cards['Physics'], False),
    ]
    for pid, iclass, icon, title, desc, cards, is_active in tab_subjects:
        panels += section_panel(pid, iclass, icon, title, f' {desc}',
                                f'stats-{pid}', f'grid-{pid}', cards, f'noresults-{pid}',
                                active=is_active)

    panels += section_panel('double', 'double', '&#128300;',
        'Double Award Science &mdash; 4DS0',
        ' Core lessons across all three sciences. B/C/P-only spec points excluded. Use the <strong>Subject</strong> filter above to jump to a science.',
        'stats-double', 'grid-double', double_html, 'noresults-double', grid_wrapper=False)

    panels += section_panel('sa', 'sa', '&#127981;',
        'Single Award Science &mdash; 4SS0',
        ' Subset of Double Award per the 4SS0 specification. Lesson numbers preserved. Use the <strong>Subject</strong> filter above to jump to a science.',
        'stats-sa', 'grid-sa', sa_html, 'noresults-sa', grid_wrapper=False)

    # PWA meta tags and manifest
    edition = 'teacher' if teacher else 'student'
    manifest_file = f'manifest_{edition}.json'
    pwa_meta = f'''<link rel="manifest" href="{manifest_file}">
<meta name="theme-color" content="{'#1a1a2e' if teacher else '#1a1a2e'}">
<meta name="apple-mobile-web-app-capable" content="yes">
<meta name="apple-mobile-web-app-status-bar-style" content="black-translucent">
<meta name="apple-mobile-web-app-title" content="{'SoW Teacher' if teacher else 'SoW Student'}">
<link rel="apple-touch-icon" href="icons/{edition}-192x192.png">
<link rel="icon" type="image/png" sizes="192x192" href="icons/{edition}-192x192.png">
<link rel="icon" type="image/png" sizes="512x512" href="icons/{edition}-512x512.png">'''

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
{pwa_meta}
<title>IGCSE Science Scheme of Work{'— Teacher Edition' if teacher else ''} | Edexcel 9-1</title>
<style>{get_css(teacher)}</style>
</head>
<body>
<header class="site-header">
  <div>
    <span class="header-eyebrow">Edexcel International GCSE &nbsp;&#183;&nbsp; 9-1 Grading</span>
    {'<span class="teacher-banner">&#128218; Teacher Edition</span> <a href="Science_Scheme_of_Work.html" class="edition-link" target="_blank">&#127891; Student Edition &rarr;</a>' if teacher else ''}
  </div>
  <h1 class="header-title">Science Scheme of Work</h1>
  <p class="header-subtitle">Lesson resource links &middot; Specification points &middot; Textbook references{teacher_note}</p>
  <div class="spec-codes">
    <a class="spec-tag" href="https://qualifications.pearson.com/en/qualifications/edexcel-international-gcses/international-gcse-biology-2017.html" target="_blank">4BI0 &mdash; Biology</a>
    <a class="spec-tag" href="https://qualifications.pearson.com/en/qualifications/edexcel-international-gcses/international-gcse-chemistry-2017.html" target="_blank">4CH0 &mdash; Chemistry</a>
    <a class="spec-tag" href="https://qualifications.pearson.com/en/qualifications/edexcel-international-gcses/international-gcse-physics-2017.html" target="_blank">4PH0 &mdash; Physics</a>
    <a class="spec-tag" href="https://qualifications.pearson.com/en/qualifications/edexcel-international-gcses/international-gcse-science-double-award-2017.html" target="_blank">4DS0 &mdash; Double Award</a>
    <a class="spec-tag" href="https://qualifications.pearson.com/en/qualifications/edexcel-international-gcses/international-gcse-science-single-award-2017.html" target="_blank">4SS0 &mdash; Single Award</a>
  </div>
  {assess_dropdown}
</header>
<nav class="nav-wrapper">
  <div class="nav-tabs">
    <button class="nav-tab active" data-tab="biology"><span class="tab-dot"></span>Biology Single Science</button>
    <button class="nav-tab" data-tab="chemistry"><span class="tab-dot"></span>Chemistry Single Science</button>
    <button class="nav-tab" data-tab="physics"><span class="tab-dot"></span>Physics Single Science</button>
    <button class="nav-tab" data-tab="double"><span class="tab-dot"></span>Double Science</button>
    <button class="nav-tab" data-tab="sa"><span class="tab-dot"></span>Single Award Science</button>
  </div>
</nav>
<div class="filter-bar">
  <div class="filter-inner">
    <span class="filter-label">Year</span>
    <button class="filter-btn active" data-year="all">All Years</button>
    <button class="filter-btn" data-year="year10">Year 10</button>
    <button class="filter-btn" data-year="year11">Year 11</button>
    <div class="filter-divider"></div>
    <span class="filter-label">Term</span>
    <button class="filter-btn active" data-term="all">All Terms</button>
    <button class="filter-btn" data-term="term1">Term 1</button>
    <button class="filter-btn" data-term="term2">Term 2</button>
    <button class="filter-btn" data-term="term3">Term 3</button>
    <div class="filter-divider"></div>
    {SUBJ_FILTER_HTML}
    <input class="search-box" type="text" placeholder="Search lessons, spec points{'or resources' if teacher else ''}&#8230;" id="searchBox">
  </div>
</div>
<main class="content-area">
  {panels}
</main>

<!-- Full-screen card modal -->
<div class="card-modal-overlay" id="cardModal" onclick="closeCardModal(event)">
  <div class="card-modal" id="cardModalInner">
    <div class="card-modal-header" id="cardModalHeader"></div>
    <button class="card-modal-close" onclick="document.getElementById('cardModal').classList.remove('open')" title="Close">&#x2715;</button>
    <div class="card-modal-body" id="cardModalBody"></div>
  </div>
</div>
<footer class="site-footer">
  <span>Updated {timestamp}</span>
</footer>
<script>{JS}</script>
<script>
if('serviceWorker' in navigator){{navigator.serviceWorker.register('sw_{edition}.js').then(function(r){{console.log('SW registered',r.scope)}}).catch(function(e){{console.log('SW failed',e)}})}}
</script>
</body>
</html>"""

# ── Main ───────────────────────────────────────────────────────────────────────
def main(force=False):
    if not XLSX_PATH.exists():
        print(f"ERROR: Spreadsheet not found at {XLSX_PATH}")
        sys.exit(1)

    # Check if spreadsheet changed since last build
    xlsx_mtime = XLSX_PATH.stat().st_mtime
    stamp_file = BASE_DIR / '.last_build_mtime'
    if not force and stamp_file.exists():
        try:
            last = float(stamp_file.read_text().strip())
            if xlsx_mtime <= last:
                print(f"No changes detected — skipping rebuild ({datetime.now().strftime('%H:%M:%S')})")
                return False
        except Exception:
            pass

    print(f"[{datetime.now().strftime('%H:%M:%S')}] Change detected — rebuilding sites...")

    if SA_SPEC_PATH.exists():
        with open(SA_SPEC_PATH) as f:
            sa_specs = json.load(f)
    else:
        sa_specs = SA_SPECS_INLINE

    sa_sets = {s: set(v) for s, v in sa_specs.items()}
    data    = load_data()

    for teacher, out_path in [(False, STUDENT_OUT), (True, TEACHER_OUT)]:
        html = build_html(data, sa_sets, teacher=teacher)
        out_path.write_text(html, encoding='utf-8')
        label = 'Teacher' if teacher else 'Student'
        print(f"  ✓ {label}: {out_path.name}  ({len(html)//1024} KB)")

    stamp_file.write_text(str(xlsx_mtime))

    # ── Generate PWA manifest and service worker files ────────────────────────
    for edition, html_file, title, short_name, color in [
        ('student', 'Science_Scheme_of_Work.html',
         'IGCSE Science Scheme of Work', 'SoW Student', '#1a1a2e'),
        ('teacher', 'Science_Scheme_of_Work_TEACHER.html',
         'IGCSE Science SoW — Teacher', 'SoW Teacher', '#1a1a2e'),
    ]:
        sizes = [72, 96, 128, 144, 152, 192, 384, 512]
        icons_list = [
            {"src": f"icons/{edition}-{s}x{s}.png", "sizes": f"{s}x{s}", "type": "image/png"}
            for s in sizes
        ]
        manifest = {
            "name": title,
            "short_name": short_name,
            "description": f"IGCSE Science Scheme of Work — {edition.title()} Edition",
            "start_url": f"./{html_file}",
            "display": "standalone",
            "background_color": "#f4f6fb",
            "theme_color": color,
            "orientation": "any",
            "icons": icons_list,
        }
        mf_path = BASE_DIR / f'manifest_{edition}.json'
        mf_path.write_text(json.dumps(manifest, indent=2), encoding='utf-8')
        print(f"  ✓ PWA manifest: {mf_path.name}")

        # Service worker — caches the HTML file for offline use
        sw_content = f"""// Service Worker for {edition.title()} Edition
const CACHE_NAME = 'sow-{edition}-v{int(time.time())}';
const URLS_TO_CACHE = ['./{html_file}'];

self.addEventListener('install', function(event) {{
  event.waitUntil(
    caches.open(CACHE_NAME).then(function(cache) {{
      return cache.addAll(URLS_TO_CACHE);
    }})
  );
  self.skipWaiting();
}});

self.addEventListener('activate', function(event) {{
  event.waitUntil(
    caches.keys().then(function(names) {{
      return Promise.all(
        names.filter(function(n) {{ return n.startsWith('sow-{edition}-') && n !== CACHE_NAME; }})
             .map(function(n) {{ return caches.delete(n); }})
      );
    }})
  );
  self.clients.claim();
}});

self.addEventListener('fetch', function(event) {{
  event.respondWith(
    caches.match(event.request).then(function(response) {{
      return response || fetch(event.request);
    }})
  );
}});
"""
        sw_path = BASE_DIR / f'sw_{edition}.js'
        sw_path.write_text(sw_content, encoding='utf-8')
        print(f"  ✓ Service worker: {sw_path.name}")

    # ── Optional Google Drive copy ────────────────────────────────────────────
    if GDRIVE_DIR:
        import shutil
        gdrive = Path(GDRIVE_DIR)
        if gdrive.exists():
            gdrive.mkdir(parents=True, exist_ok=True)
            for src in [STUDENT_OUT, TEACHER_OUT]:
                dest = gdrive / src.name
                shutil.copy2(src, dest)
                print(f"  ✓ Copied to Google Drive: {dest}")
        else:
            print(f"  ⚠ Google Drive folder not found: {GDRIVE_DIR} — skipping copy")

    print(f"  Done.")
    return True

if __name__ == '__main__':
    force = '--force' in sys.argv
    main(force=force)
